# excel files can also be styled using the style module it provides

import openpyxl
from pathlib import Path
from openpyxl.styles import Font, Color


sales_wb = openpyxl.load_workbook(path / 'sample-data' / 'produceSales.xlsx')
active_sheet = sales_wb.active

for row in range(2, int(active_sheet.max_row) + 1):
    produce_name = active_sheet.cell(row=row, column=1).value.lower()
    if produce_name in price_changes.keys():
        active_sheet['B'+str(row)] = price_changes[produce_name]


# say I want to style the prices of produce with prices higher than $10 and also give them a color blue
# this can be achieved using the Font and Color class

font_color = Color('0000ff')
font = Font(family='Times New Roman', size=14, bold=True, italic=True, color=font_color)

active_sheet['A10'].font = font


sales_wb.save(path / 'sample-data' / 'produceSales.xlsx')
