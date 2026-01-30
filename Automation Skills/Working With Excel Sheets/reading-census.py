# In this exercise, I will be reading and Excel sheet and provide a stat about the data in the file
# the stats that will be provided will be, the number of tracts in each country and the total population in each country
import openpyxl
from pathlib import Path

file_path = Path(__file__).parent
wb = openpyxl.load_workbook(filename=file_path / 'sample-data'/'censuspopdata.xlsx')

sheet1 = wb.active

print("Size of the sheet is {sheet1.max_row} X {sheet1.max_column}".format(sheet1=sheet1))

data = {}

# looping through the sheet
for row in range(2, int(sheet1.max_row)+1):
    state = sheet1['B' + str(row)].value
    country = sheet1['C' + str(row)].value
    population = int(sheet1['D' +str(row)].value)

    data.setdefault(state, {})
    data[state].setdefault(country, {})

    data[state][country]['pop'] = data[state][country].get('pop', 0) + population
    data[state][country]['tract'] = data[state][country].get('tract', 0) + 1

# and now I can get the tract and population for each country as long as the state name is known
print(data['AK']['Anchorage']['pop'])
