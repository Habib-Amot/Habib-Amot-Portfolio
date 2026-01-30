# A workbook can be created by creating an instance of The Workbook class

import openpyxl
from pathlib import Path

path = Path(__file__).parent

wb = openpyxl.Workbook()

# create a sheet in the workbook
first_sheet = wb.create_sheet('business leads', index=0)

# getting the sheet names
print(wb.sheetnames)

# delete the second sheet
del wb['Sheet']

# adding some data into the spreadsheet
first_sheet['A1'] = 'Date Entered'
first_sheet['B1'] = 'Company Name'
first_sheet['C1'] = 'Company Email Address'
first_sheet['D1'] = 'Email sent'


# and the workbook can be saved
wb.save(filename=path / 'business-leads.xlsx')

# say we got a price increase for our products which were recorded in an EXcel file and now wants to change the price to the current one

price_changes = {
    'garlic': 0.5,
    'celery': 1.3,
    'onions': 0.05
}

sales_wb = openpyxl.load_workbook(path / 'sample-data' / 'produceSales.xlsx')
active_sheet = sales_wb.active

for row in range(2, int(active_sheet.max_row) + 1):
    produce_name = active_sheet.cell(row=row, column=1).value.lower()
    if produce_name in price_changes.keys():
        active_sheet['B'+str(row)] = price_changes[produce_name]

sales_wb.save(path / 'sample-data' / 'produceSales.xlsx')
