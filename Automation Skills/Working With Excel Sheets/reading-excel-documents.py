import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string

file_path = Path(__file__).parent

""" path = Path(__file__).parent

with zipfile.ZipFile(path / 'Automate_the_Boring_Stuff_2e_onlinematerials.zip') as zfile:
    zfile.extract('automate_online-materials/example.xlsx', path=path)
    zfile.extract('automate_online-materials/produceSales.xlsx', path=path)
    [print(file.filename) for file in zfile.filelist if file.filename.endswith('.xlsx')] """

# In Excel a workbook is a collection of sheets and each sheet is a collection of rows and columns
# to work with individual sheets in a workbook, the workbook needs to be loaded first 

wb = openpyxl.load_workbook(filename=file_path / 'sample-data'/'example.xlsx')

# and now, the worksheets in the workbook can now be accessed
# first, we get all the worksheets that are present in the workbook
all_sheets = wb.sheetnames
print("this are all the sheets that are available in the workbook example.xlsx", all_sheets)

# accessing individual worksheet
sheet1 = wb['Sheet1']

# the size of a sheet can be gotten by accessing the max_column and max_row attribute
print("sheet1 size is {sheet1.max_row} X {sheet1.max_column}".format(sheet1=sheet1))
# and now the individual cells, rows and columns can be accessed

''' NB: 
        Rows are labeled from 1
        While columns are labeled with Alphabets from A
'''

column_names = [get_column_letter(col_number) for col_number in range(1, sheet1.max_column+1)]
print("Available columns in sheet1", " ".join(column_names))

print("\n{0:=^50}".format("Accessing Cells in a Sheet"))
# accessing individual sheet

cell_A1 = sheet1['A1']
cell_A2 = sheet1['A2']

print("the value of cell A1 is", cell_A1.value)
print("the value of cell A2 is", cell_A2.value)

print()
# and also a range of cells can be indexed
cell_range = sheet1["A1:C3"] # this does not just return a collections of Cells, but a colllection of cells grouped in tuple-rows
for row_of_cells in cell_range:
    for cells in row_of_cells:
        print(cells.value, end=" ")
    print()

print("\n{0:=^60}".format("Converting Between Column Letters and Numbers"))

# since column names can be very trickish to work with sometimes, the get_column_letter and column_index_from_string function are
# available to convert between column index to name and column name to index respectively

print("The equivalent number for column letter ACH is", column_index_from_string("ACH"))
print("The equivalent letter for column number 100 is", get_column_letter(1000))
